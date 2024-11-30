import itertools
import math
import statistics as stats
import pandas as pd
import numpy as np
import xlsxwriter
from os.path import exists
from openpyxl import Workbook
from openpyxl import load_workbook

def nucleiAve(nuclei_list): #return average num nuclei
    nuclei = []
    for i in range(len(nuclei_list)):
        if nuclei_list[i] == 'CELL' or i == len(nuclei_list) - 1:
            if i != len(nuclei_list) - 1:
                nuclei.append(float(nuclei_list[i - 1]) - 2)
            else:
                nuclei.append(float(nuclei_list[i]) - 2)
    return len(nuclei), stats.mean(nuclei), nuclei


def minInterDist(x, y, n): #returns the minimum internuclear distance from nuclei n
    dists = []
    for i in range(len(x)):
        if i != n:
            x_val = (float(x[n]) - float(x[i]))**2
            y_val = (float(y[n]) - float(y[i]))**2
            dist = math.sqrt(x_val + y_val)
            dists.append(dist)
    return min(dists)

def interDist(x, y): #return average and standard dev of internuclear distance
    ave_inter_dists = []
    min_inter_dists = []
    total_inter_dists = []
    break_points = []
    break_points.append(0)
    for i in range(len(x)):
        if x[i] == 'CELL':
            break_points.append(i + 1)
    break_points.append(len(x) + 1)
    count = 0
    for i in range(len(x) + 1):
        if i == len(x) or x[i] == 'CELL':
            ave_inter_dists.append(stats.mean(min_inter_dists))
            total_inter_dists.append(min_inter_dists)
            min_inter_dists = []
            count += 1
            continue
        else:
            x_list = x[break_points[count] : break_points[count+1] - 1]
            y_list = y[break_points[count] : break_points[count+1] - 1]
            min_inter_dists.append(minInterDist(x_list, y_list, i - break_points[count])) 
    return stats.mean(ave_inter_dists), stats.stdev(ave_inter_dists), total_inter_dists


def muscleLen(muscles): #return average muscle length (in microns too)
    muscle_lens = []
    for val in muscles:
        if val != 'CELL' and float(val) > 1:
            muscle_lens.append(float(val))
    ave = stats.mean(muscle_lens)
    return ave, muscle_lens


def muscleArea(areas): #return average muscles area (in microns too)
    muscle_areas = []
    for val in areas:
        if val != 'CELL':
            muscle_areas.append(float(val))
    ave = stats.mean(muscle_areas)
    return ave, muscle_areas

def idealAreaDist(num_nuclei, area): #returns ideal area and distance for one muscle CELL
    ideal_area = area/num_nuclei
    ideal_dist = ideal_area**0.5
    return ideal_area, ideal_dist

def aiRatio(nuclei, areas, dists): #return the average A:I ratio
    ideal_a = []
    ideal_d = []
    aiRatios = []
    count = 0
    for i in range(len(areas)):
        if areas[i] != 'CELL':
            a, d = idealAreaDist(int(nuclei[count]), float(areas[i]))
            ideal_a.append(a)
            ideal_d.append(d)
            count += 1
    count = 0
    for l in dists:
        for val in l:
            aiRatios.append(val/ideal_d[count])
        count += 1
    return stats.mean(aiRatios), stats.stdev(aiRatios), aiRatios

def nucleiToEdge_distance(nuclei_to_edge):
    cell_ave_distance = []
    distances = []
    break_point = [0]
    for i, val in enumerate(nuclei_to_edge):
        if val != 'CELL':
            distances.append(val)
        else:
            break_point.append(len(distances))
    break_point.append(len(distances))
    for i in range(len(break_point) - 1):
        cell_ave_distance.append(stats.mean(distances[break_point[i] : break_point[i+1]]))
    break_point.pop()
    larvae_average = stats.mean(cell_ave_distance)
    return cell_ave_distance, break_point, distances, larvae_average

def cov_calculator(area):
    mean = stats.mean(area)
    std = stats.stdev(area)
    return std/mean * 100

def voronoi_variability(voronoi_area):
    cov = []
    polygon_area = []
    break_point = [0]
    ave_polygon_area = []
    for val in voronoi_area:
        if val != 'CELL':
            polygon_area.append(val)
        else:
            break_point.append(len(polygon_area))
    break_point.append(len(polygon_area))
    for i in range(len(break_point) - 1):
        ave_polygon_area.append(stats.mean(polygon_area[break_point[i] : break_point[i+1]]))
        cov.append(cov_calculator(polygon_area[break_point[i] : break_point[i+1]]))
    break_point.pop()
    larvae_average = stats.mean(ave_polygon_area)

    return cov, break_point, polygon_area, ave_polygon_area, larvae_average
        

def main(nuclei, area, x, y, length, nuclei_to_edge, voronoi_area, wb_name, ws_name, iteration):
        
    #The lists are already formatted correctly and can be passed in for analysis 
    num_muscles, n_ave, n_list = nucleiAve(nuclei)
    id_ave, id_std, dists = interDist(x, y)
    ave_m_len, m_lens = muscleLen(length)
    ave_m_area, m_areas = muscleArea(area)
    ave_ai_rat, std_ai_rat, ratios = aiRatio(n_list, area, dists)
    ave_distances, break_point, total_distances, larvae_ave_nuclei_to_edge = nucleiToEdge_distance(nuclei_to_edge)
    cov, cov_break_point, total_voronoi_area, cell_ave_area, larvae_ave_area = voronoi_variability(voronoi_area)

    file_exists = exists(wb_name)
    if not file_exists:
        wb = Workbook()
    else:
        wb = load_workbook(wb_name)
    ws = wb.create_sheet(ws_name)

    c_row = 1
    m_count = 1
    for i in range(len(n_list)):
        ws['A' + str(c_row)] = 'Muscle: ' + str(iteration + 1) + '.' + str(m_count)

        ws['B' + str(c_row)] = 'Number of nuclei'
        ws['B' + str(c_row + 1)] = n_list[i]

        ws['C' + str(c_row)] = 'Internuclear distances'
        j = 0
        for x in range(c_row, c_row + len(dists[i])):
            ws['C' + str(x + 1)] = dists[i][j]
            j += 1
        ws['D' + str(c_row)] = 'Average Internuclear distances'
        ws['D' + str(c_row + 1)] = stats.mean(dists[i])

        ws['E' + str(c_row)] = 'Std internuclear distance'
        ws['E' + str(c_row + 1)] = stats.stdev(dists[i])

        ws['F' + str(c_row)] = 'Muscle length (micron)'
        ws['F' + str(c_row + 1)] = m_lens[i]

        ws['G' + str(c_row)] = 'Muscle area (micron)'
        ws['G' + str(c_row + 1)] = m_areas[i]

        ws['H' + str(c_row)] = 'A:I ratios'
        m_ratios = []
        for x in range(c_row, c_row + len(dists[i])):
            ws['H' + str(x + 1)] = ratios[0]
            m_ratios.append(ratios[0])
            ratios.pop(0)
                
        ws['I' + str(c_row)] = 'Muscle average A:I ratio'
        ws['I' + str(c_row + 1)] = stats.mean(m_ratios)
                
        ws['J' + str(c_row)] = 'Muscle std A:I ratio'
        ws['J' + str(c_row + 1)] = stats.stdev(m_ratios)
        
        ws['K' + str(c_row)] = 'Voronoi Area COV'
        ws['K' + str(c_row + 1)] = cov[i]

        ws['L' + str(c_row)] = 'Voronoi Area (micron)'
        k = 0
        for x in range(c_row, c_row + len(dists[i])):
            ws['L' + str(x+1)] = total_voronoi_area[cov_break_point[i] + k]
            k += 1

        ws['M' + str(c_row)] = 'Average Voronoi Area (micron)'
        ws['M' + str(c_row + 1)] = cell_ave_area[i]

        ws['N' + str(c_row)] = 'Nuclei to Edge Distance'
        j = 0
        for x in range(c_row, c_row + len(dists[i])):
            ws['N' + str(x + 1)] = total_distances[break_point[i] + j]
            j += 1
            
        ws['O' + str(c_row)] = 'Average Nuclei to Edge Distances'
        ws['O' + str(c_row + 1)] = ave_distances[i]
        
        m_ratios = []
        c_row = c_row + len(dists[i]) + 3
        m_count += 1
        


    ws['P1'] = 'Larva Averages'
    
    ws['Q1'] = 'Average nuclei'
    ws['Q2'] = n_ave

    ws['R1'] = 'Average internuclear distance'
    ws['R2'] = id_ave

    ws['S1'] = 'Standard dev internuclear distance'
    ws['S2'] = id_std

    ws['T1'] = 'Average muscle length (micron)'
    ws['T2'] = ave_m_len

    ws['U1'] = 'Average muscle area (micron)'
    ws['U2'] = ave_m_area

    ws['V1'] = 'Average A:I ratio'
    ws['V2'] = ave_ai_rat

    ws['W1'] = 'Standard deviation A:I ratio'
    ws['W2'] = std_ai_rat

    ws['X1'] = 'Average Voronoi Are COV'
    ws['X2'] = stats.mean(cov)

    ws['Y1'] = 'Average Voronoi Area'
    ws['Y2'] = larvae_ave_area
    
    ws['Z1'] = 'Average Nuclei to Edge Distance (micron)'
    ws['Z2'] = larvae_ave_nuclei_to_edge


    wb.save(wb_name)
